#!/usr/bin/env python3
"""Genera una app web autocontenida desde la hoja DATOS de una planilla.

Produce un único .html que abre en cualquier celular sin instalar nada: muestra
la sesión del día, permite registrar reps, kilos y RIR, y exporta lo cargado a
CSV. Es la solución puente hasta que el frontend esté listo.

Uso:
    python scripts/gen_app.py ../data/planilla.xlsx rutina.html
"""

from __future__ import annotations

import argparse
import collections
import json
from pathlib import Path
from typing import Any

import openpyxl

TEMPLATE = Path(__file__).resolve().parent / "template.html"
PLACEHOLDER = "/*__DATA__*/null"


def read_workbook(xlsx: Path) -> dict[str, Any]:
    """Aplana la hoja DATOS a la estructura que consume el template."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["DATOS"]

    header = [cell.value for cell in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    sheet = wb["ATLETA"]
    athlete = sheet["B5"].value or "Atleta"
    mesocycles: dict[str, str] = {}
    for row in range(14, 22):
        number = sheet.cell(row=row, column=1).value
        label = sheet.cell(row=row, column=2).value
        if number and label:
            mesocycles[str(int(number))] = label

    sessions: collections.OrderedDict[tuple, collections.OrderedDict] = collections.OrderedDict()
    for raw in ws.iter_rows(min_row=2, values_only=True):
        name = raw[idx["Ejercicio"]]
        if not name:
            continue

        key = (raw[idx["Semana"]], raw[idx["Sesión"]])
        block = sessions.setdefault(key, collections.OrderedDict())
        exercise = block.setdefault(
            name,
            {
                "nombre": name,
                "patron": raw[idx["Patrón"]],
                "descanso": raw[idx["Descanso"]],
                "obs": raw[idx["Observación"]],
                "meso": raw[idx["Meso #"]],
                "series": [],
            },
        )
        # El descanso y la observación se cargan a nivel ejercicio en la planilla,
        # pero sólo en la fila de la primera serie.
        if raw[idx["Descanso"]] and not exercise["descanso"]:
            exercise["descanso"] = raw[idx["Descanso"]]
        if raw[idx["Observación"]] and not exercise["obs"]:
            exercise["obs"] = raw[idx["Observación"]]

        exercise["series"].append(
            {
                "n": raw[idx["Serie #"]],
                "rmin": raw[idx["Reps plan mín"]],
                "rmax": raw[idx["Reps plan máx"]],
                "rirmin": raw[idx["RIR plan mín"]],
                "rirmax": raw[idx["RIR plan máx"]],
                "kg": raw[idx["Kg plan"]],
                "reps": raw[idx["Reps real"]],
                "kgr": raw[idx["Kg real"]],
                "rir": raw[idx["RIR real"]],
            }
        )

    out = []
    for (week, day), exercises in sessions.items():
        items = list(exercises.values())
        out.append(
            {
                "semana": week,
                "sesion": day,
                "meso": items[0]["meso"] if items else None,
                "ejercicios": items,
            }
        )
    out.sort(key=lambda s: (s["semana"], s["sesion"] or 0))

    return {"atleta": athlete, "mesos": mesocycles, "sesiones": out}


def render(data: dict[str, Any], template: Path) -> str:
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    # Un "</" dentro del JSON cerraría el <script> que lo contiene.
    payload = payload.replace("</", "<\\/")
    return template.read_text(encoding="utf-8").replace(PLACEHOLDER, payload)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="planilla de origen")
    parser.add_argument(
        "output", type=Path, nargs="?", default=Path("rutina.html"), help="html de salida"
    )
    parser.add_argument("--template", type=Path, default=TEMPLATE)
    args = parser.parse_args()

    if not args.template.exists():
        raise SystemExit(f"falta el template: {args.template}")

    data = read_workbook(args.xlsx)
    args.output.write_text(render(data, args.template), encoding="utf-8")

    exercises = sum(len(s["ejercicios"]) for s in data["sesiones"])
    size_kb = args.output.stat().st_size // 1024
    print(f"{args.output}: {len(data['sesiones'])} sesiones, {exercises} ejercicios, {size_kb} KB")


if __name__ == "__main__":
    main()
