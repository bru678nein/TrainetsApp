"""Import the training spreadsheet into the relational schema.

Usage:
    python -m importer.from_spreadsheet ../data/planilla.xlsx [--reset]

The database must already be migrated (`alembic upgrade head`). This script
populates, it does not create schema: it used to run `drop_all` + `create_all`,
which produced a schema parallel to the migrations — no view, no functional
index, no extensions — and on top of that wiped whatever database it was handed.

It exists so development runs on real data from day 1 instead of invented seeds.
"""

from __future__ import annotations

import os
import re
import sys
import unicodedata
from collections import defaultdict
from decimal import Decimal
from typing import Any

import openpyxl
from sqlalchemy import create_engine, func, select, text
from sqlalchemy.orm import Session as OrmSession

from app.domain.rpe import OutOfChartError, estimate_1rm
from app.models import (
    AppUser,
    Athlete,
    Coach,
    Exercise,
    LoggedSet,
    Mesocycle,
    MovementPattern,
    PrescribedSet,
    Prescription,
    Program,
    Session,
)

# Reverse dependency order is unnecessary: TRUNCATE ... CASCADE handles it,
# but listing them explicitly documents the blast radius of the delete.
#
# app_user has to be here even though CASCADE from coach does not reach it: the
# foreign key points the other way, coach -> app_user. Leaving it out let the
# identity survive a --reset, and the next seed hit the UNIQUE on auth_user_id.
#
# movement_pattern is not listed, and that changes nothing on its own: TRUNCATE
# ... CASCADE empties every table holding a foreign key into the named ones, and
# since 0018 movement_pattern.coach_id points at coach. Measured — dropping it
# from this list still left zero rows. The shared base is restored explicitly
# below instead.
_SEEDED_TABLES = "app_user, coach, exercise"


def borrar_conservando_la_base(db: OrmSession) -> None:
    """What `--reset` deletes, minus the shared pattern vocabulary.

    Reading the rows out and putting them back is not belt and braces. TRUNCATE
    ... CASCADE empties every table with a foreign key into the named ones, and
    since 0018 `movement_pattern.coach_id` points at `coach` — so the base goes
    whether or not the table is listed. Measured, not assumed.

    Losing it is not recoverable by re-running anything: migration 0019 puts the
    eleven there and no migration runs twice. The database would be left where
    production was before 0019 — unable to create a single exercise, because
    `exercise.pattern_code` is NOT NULL and references this table.

    It restores whatever the database held as shared, never a hardcoded copy of
    the eleven. The migration stays the only place that decides what the base is.
    """
    base = [
        dict(fila._mapping)
        for fila in db.execute(
            text(
                "SELECT code, label_es, is_compound, sort_order "
                "FROM movement_pattern WHERE coach_id IS NULL"
            )
        )
    ]
    db.execute(text(f"TRUNCATE {_SEEDED_TABLES} RESTART IDENTITY CASCADE"))
    if base:
        db.execute(
            text(
                "INSERT INTO movement_pattern (code, label_es, is_compound, sort_order) "
                "VALUES (:code, :label_es, :is_compound, :sort_order)"
            ),
            base,
        )


def slug(raw: str) -> str:
    n = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", n.lower()).strip("_")[:40]


def dec(value: Any) -> Decimal | None:
    """`numeric` columns map to `Decimal`.

    openpyxl returns `float`, and handing that straight to psycopg sends it as
    float8 for Postgres to round. Converting via `str` avoids dragging the
    float's binary noise along.
    """
    if value is None or value == "":
        return None
    return Decimal(str(value))


def clean_range(lo: Any, hi: Any, label: str, flags: list[str]) -> tuple[int | None, int | None]:
    """An inverted range means the original text was a compound prescription
    ("8 a 12 + 2x 3 a 5") that the parser cannot disambiguate. No value is
    invented: it is left null and flagged for review."""
    if lo is not None and hi is not None and hi < lo:
        flags.append(label)
        return None, None
    return lo, hi


def read_rows(xlsx: str) -> tuple[list[dict[str, Any]], str, dict[int, str]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["DATOS"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[idx["Ejercicio"]]:
            continue
        rows.append({k: raw[i] for k, i in idx.items()})
    at = wb["ATLETA"]
    athlete_name = at["B5"].value or "Atleta"
    mesos = {}
    for r in range(14, 22):
        n, label = at.cell(row=r, column=1).value, at.cell(row=r, column=2).value
        if n and label:
            mesos[int(n)] = label
    return rows, athlete_name, mesos


def construir_estructura(
    db: OrmSession,
    coach: Coach,
    athlete: Athlete,
    rows: list[dict[str, Any]],
    meso_labels: dict[int, str],
    program_name: str,
    stats: defaultdict[str, int],
    review: list[str],
    con_registros: bool,
) -> None:
    """Lo que la planilla describe, como filas de la base.

    Sale de `run()` para que lo pueda usar también el endpoint de importación,
    que corre **dentro del contexto del entrenador** y no como dueño. Los dos
    caminos construyen lo mismo, o la beta importaría algo distinto de lo que
    importó la migración original — que es justo lo que no se quiere.

    `con_registros` es la única diferencia, y no es una opción de conveniencia:
    bajo RLS un entrenador **no puede escribir `logged_set`** —la policy lo
    rechaza, medido contra la base— porque registrar es el acto del atleta. El
    script corre como dueño y sí puede; la API no, y no debería.
    """
    patterns: dict[str, MovementPattern] = {}
    for i, label in enumerate(sorted({r["Patrón"] for r in rows if r["Patrón"]})):
        code = slug(label)
        mp = db.get(MovementPattern, code)
        if mp is None:
            mp = MovementPattern(code=code, label_es=label, sort_order=i)
            db.add(mp)
            stats["patterns"] += 1
        patterns[label] = mp

    # Reusar antes de crear, igual que con los patrones de arriba, y por un motivo
    # que sólo aparece importando sobre un espacio que ya se usó: el catálogo
    # tiene un único por (entrenador, nombre). Un entrenador que ya venía
    # trabajando en la aplicación tiene la mitad de estos ejercicios cargados, y
    # crearlos a ciegas revienta la importación entera contra ese índice.
    #
    # `run()` nunca lo vio porque arranca de una base vacía. Lo encontró el test
    # del endpoint, que importa sobre el espacio sembrado.
    # La clave es el nombre **en minúsculas**, igual que el índice.
    #
    # `exercise_name_scope_idx` es único sobre `(coach_id, lower(name))`, así que
    # «REMO AL MENTÓN» y «Remo al mentón» son el mismo ejercicio para la base y
    # dos distintos para un `==`. La plantilla los escribe de las dos formas en
    # hojas distintas, y sin esto la importación entera muere en la primera
    # repetición. Encontrado corriéndolo contra el archivo real, no leyéndolo.
    exercises: dict[str, Exercise] = {}
    for r in rows:
        name = r["Ejercicio"]
        clave = str(name).strip().lower()
        if clave in exercises:
            exercises[name] = exercises[clave]
            continue
        ex = db.scalars(
            select(Exercise).where(
                Exercise.coach_id == coach.id, func.lower(Exercise.name) == clave
            )
        ).first()
        if ex is None:
            ex = Exercise(
                coach=coach,
                pattern_code=patterns[r["Patrón"]].code,
                name=name,
                is_competition_lift=(r["Básico"] == "Sí"),
            )
            db.add(ex)
            stats["exercises"] += 1
        exercises[clave] = ex
        exercises[name] = ex

    program = Program(coach=coach, athlete=athlete, name=program_name, status="completed")
    db.add(program)

    # Weeks per mesocycle, for week_count and for the relative week number
    weeks_by_meso: dict[int, set[int]] = defaultdict(set)
    for r in rows:
        weeks_by_meso[int(r["Meso #"])].add(int(r["Semana"]))

    mesos: dict[int, Mesocycle] = {}
    for n in sorted(weeks_by_meso):
        m = Mesocycle(
            program=program,
            ordinal=n,
            week_count=len(weeks_by_meso[n]),
            label=meso_labels.get(n, f"Mesociclo {n}"),
        )
        mesos[n] = m
        db.add(m)
        stats["mesocycles"] += 1

    sessions: dict[tuple[int, int], Session] = {}
    prescriptions: dict[tuple[int, int, str], Prescription] = {}

    for r in rows:
        meso_n, week_g, day = int(r["Meso #"]), int(r["Semana"]), int(r["Sesión"] or 1)
        # Week number relative to the mesocycle
        week_rel = sorted(weeks_by_meso[meso_n]).index(week_g) + 1
        skey = (week_g, day)
        if skey not in sessions:
            sessions[skey] = Session(mesocycle=mesos[meso_n], week_number=week_rel, day_number=day)
            db.add(sessions[skey])
            stats["sessions"] += 1

        pkey = (week_g, day, r["Ejercicio"])
        if pkey not in prescriptions:
            pos = len([k for k in prescriptions if k[0] == week_g and k[1] == day]) + 1
            rest = r["Descanso"]
            secs = None
            if isinstance(rest, str):
                nums = re.findall(r"\d+", rest)
                if nums:
                    secs = int(nums[-1]) * 60
            prescriptions[pkey] = Prescription(
                session=sessions[skey],
                exercise=exercises[r["Ejercicio"]],
                position=pos,
                rest_seconds=secs,
                coach_note=r["Observación"],
            )
            db.add(prescriptions[pkey])
            stats["prescriptions"] += 1

        rmin, rmax = clean_range(
            r["Reps plan mín"], r["Reps plan máx"], f"S{week_g} D{day} {r['Ejercicio']}", review
        )
        ps = PrescribedSet(
            prescription=prescriptions[pkey],
            set_number=int(r["Serie #"]),
            reps_min=rmin,
            reps_max=rmax,
            rir_min=dec(r["RIR plan mín"]),
            rir_max=dec(r["RIR plan máx"]),
            target_load_kg=dec(r["Kg plan"]),
        )
        db.add(ps)
        stats["prescribed_sets"] += 1

        if con_registros and r["Reps real"] is not None:
            e1rm = None
            if r["Kg real"] and r["RIR real"] is not None:
                try:
                    e1rm = estimate_1rm(
                        float(r["Kg real"]), int(r["Reps real"]), float(r["RIR real"])
                    )
                except (OutOfChartError, ValueError):
                    stats["e1rm_out_of_chart"] += 1
            db.add(
                LoggedSet(
                    prescribed_set=ps,
                    athlete=athlete,
                    reps=r["Reps real"],
                    load_kg=dec(r["Kg real"]),
                    rir=dec(r["RIR real"]),
                    athlete_note=r["Comentario"],
                    e1rm_kg=dec(e1rm),
                )
            )
            stats["logged_sets"] += 1


def run(xlsx: str, dsn: str, reset: bool = False) -> tuple[dict[str, int], list[str]]:
    """Populate an already-migrated database. Returns (counts, sets needing review)."""
    rows, athlete_name, meso_labels = read_rows(xlsx)
    engine = create_engine(dsn)
    stats: defaultdict[str, int] = defaultdict(int)
    review: list[str] = []

    with OrmSession(engine) as db:
        # Both, not just Coach: an identity can outlive its coach row, and then
        # the check would pass and the insert would fail on the UNIQUE instead
        # of on this message.
        if db.scalar(select(func.count()).select_from(Coach)) or db.scalar(
            select(func.count()).select_from(AppUser)
        ):
            if not reset:
                raise SystemExit(
                    "La base ya tiene datos. Pasá --reset para borrarlos, o apuntá a "
                    "otra base. No borro nada sin que me lo pidan."
                )
            borrar_conservando_la_base(db)
            db.commit()

        user = AppUser(auth_user_id="seed-coach", email="coach@example.com", display_name="Coach")
        coach = Coach(user=user)
        # user_id stays NULL: the seeded athlete has no account. That is the
        # normal state until the invitation flow fills it in.
        athlete = Athlete(coach=coach, full_name=athlete_name, level="intermedio")
        db.add_all([user, coach, athlete])

        # Reuse before creating. The shared base ships in migration 0019, so on
        # any migrated database these eleven already exist; inserting them blind
        # hits the primary key. They also survive `--reset` — they belong to the
        # schema, not to this import — so this branch is the normal path, not
        # the edge case.
        construir_estructura(
            db,
            coach,
            athlete,
            rows,
            meso_labels,
            "Migrado desde planilla",
            stats,
            review,
            con_registros=True,
        )
        db.commit()
    stats["sets_needing_review"] = len(review)
    return dict(stats), sorted(set(review))


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--reset"]
    xlsx = args[0] if args else "../data/planilla.xlsx"
    dsn = args[1] if len(args) > 1 else os.environ.get("DATABASE_URL", "")
    if not dsn:
        raise SystemExit("Falta DATABASE_URL (o pasala como segundo argumento).")
    stats, review = run(xlsx, dsn, reset="--reset" in sys.argv)
    for k, v in stats.items():
        print(f"{k:22s} {v}")
    if review:
        print("\nSeries con rango de reps ambiguo (revisar a mano):")
        for r in review:
            print("  -", r)
