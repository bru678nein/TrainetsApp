"""La plantilla que usan los entrenadores, normalizada a la forma que ya se sabe leer.

El otro lector, `from_spreadsheet.read_rows`, lee **la planilla migrada**: una
hoja `DATOS` con una fila por serie. Ésta es distinta: una hoja por bloque, y
adentro bloques de doce filas que se repiten —`Semana N`, `Sesión N`, un
encabezado, y los ejercicios abajo—.

No se escribe un segundo constructor. Este módulo devuelve exactamente la misma
tupla que `read_rows`, así que `construir_estructura` no se entera de cuál de los
dos lo alimentó. Dos lectores y un constructor: la forma de la base se decide en
un solo lugar.

**Lo que este módulo NO hace es adivinar.** Lo que no puede leer con certeza
queda en nulo y se marca, igual que `clean_range`. Un importador que completa
huecos deja prescripciones plausibles y equivocadas, y después alguien programa
arriba de eso.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any

import openpyxl

#: Las columnas, por su encabezado. Se buscan por nombre y no por posición: la
#: plantilla tiene columnas vacías entre medio y alguna hoja las corrió.
ENCABEZADO = "Patron mov."

#: Cuánto vale cada columna en la forma que espera el constructor.
COLUMNAS = {
    "Patron mov.": "Patrón",
    "Ejercicio": "Ejercicio",
    "Series": "Series",
    "Repeticiones @RPE": "Reps",
    "RIR": "RIR",
    "Kilos": "Kg plan",
    "Descanso": "Descanso",
    "Observaciones": "Observación",
}


def _numeros(texto: Any) -> list[float]:
    """Los números de «10 a 15», «@2-3» o «35k». Sin inventar los que no están."""
    if texto is None:
        return []
    if isinstance(texto, int | float) and not isinstance(texto, bool):
        return [float(texto)]
    return [float(n) for n in re.findall(r"\d+(?:[.,]\d+)?", str(texto).replace(",", "."))]


def _rango(texto: Any, donde: str, revisar: list[str]) -> tuple[float | None, float | None]:
    """Un rango, o nada.

    **Una fecha es lo que Excel hizo con lo que el entrenador escribió.** Al
    tipear `4-5` en una celda con formato general, Excel guarda el 4 de mayo. La
    intención original no se puede recuperar —`4-5` pudo ser «4 a 5» o «5 a 4»— y
    reconstruirla desde el mes y el día sería exactamente el tipo de adivinanza
    que produce una prescripción que nadie escribió.
    """
    if isinstance(texto, datetime | date):
        revisar.append(f"{donde}: Excel convirtió el valor en una fecha")
        return None, None
    nums = _numeros(texto)
    if not nums:
        return None, None
    if len(nums) == 1:
        return nums[0], nums[0]
    return min(nums[0], nums[1]), max(nums[0], nums[1])


def _patron(fila: list[Any], idx: dict[str, int]) -> str:
    """El patrón de movimiento, o uno de descarte.

    La plantilla lo deja vacío en los accesorios —vuelos laterales, gemelos— y
    esos ejercicios igual tienen que entrar: el volumen por patrón se calcula
    sobre lo que hay, y perder una fila lo cambia.
    """
    if "Patrón" not in idx:
        return "Sin patrón"
    crudo = fila[idx["Patrón"]]
    return str(crudo).strip() if crudo and str(crudo).strip() else "Sin patrón"


def leer(xlsx: str) -> tuple[list[dict[str, Any]], str, dict[int, str], list[str]]:
    """Las filas de la plantilla, con la forma de `read_rows`.

    **La hoja es una grilla, no una lista**, y eso es lo único que hay que
    entender de este formato:

    - Las **columnas** son semanas. El encabezado se repite hacia la derecha —en
      este archivo cuatro veces, en las columnas 1, 15, 29 y 43— y arriba de cada
      repetición dice `Semana 1`, `Semana 2`, `Semana 3`, `Semana 4`.
    - Las **filas** son sesiones. El mismo encabezado se repite hacia abajo, y
      arriba de cada repetición dice `Sesión N`.
    - Cada **hoja** es un bloque, y su orden es el orden de los bloques: medido,
      el mismo ejercicio baja repeticiones y sube carga de una hoja a la
      siguiente.

    Leer sólo la primera columna —que es lo que parece una lista— trae un cuarto
    del archivo y hace creer que el bloque tiene una sola semana. Es exactamente
    lo que pasó la primera vez que se escribió esto.

    Devuelve también **los avisos**, y ése es el cuarto elemento de la tupla y no
    un detalle: la primera versión los juntaba en una lista local y no la
    devolvía, así que las celdas que Excel había arruinado se perdían en silencio
    y la importación decía «nada para revisar» sobre un archivo que sí tenía.
    Un aviso que no llega es peor que no tenerlo: promete que se miró.
    """
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    filas: list[dict[str, Any]] = []
    etiquetas: dict[int, str] = {}
    revisar: list[str] = []
    bloque = 0

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        celdas = [
            [str(c).strip() if c is not None else "" for c in fila]
            for fila in ws.iter_rows(values_only=True)
        ]
        crudas = list(ws.iter_rows(values_only=True))
        encabezados = [n for n, f in enumerate(celdas) if ENCABEZADO in f]
        if not encabezados:
            continue

        bloque += 1
        etiquetas[bloque] = hoja.strip()

        for orden, n in enumerate(encabezados, 1):
            fila = celdas[n]
            # Cada aparición del encabezado a lo ancho es una semana distinta.
            for inicio in [i for i, texto in enumerate(fila) if texto == ENCABEZADO]:
                idx = {
                    COLUMNAS[texto]: inicio + d
                    for d, texto in enumerate(fila[inicio : inicio + len(COLUMNAS) + 2])
                    if texto in COLUMNAS
                }
                semana = _numero_de(celdas, n - 2, inicio, "semana") or 1
                sesion = _numero_de(celdas, n - 1, inicio, "sesión") or orden

                for cruda in crudas[n + 1 :]:
                    texto_fila = [str(c).strip() if c is not None else "" for c in cruda]
                    # El siguiente encabezado hacia abajo cierra esta sesión.
                    if ENCABEZADO in texto_fila:
                        break
                    filas.extend(_series_de(cruda, idx, bloque, semana, sesion, hoja, revisar))

    # La semana pasa a ser **global** antes de devolver.
    #
    # `construir_estructura` identifica una prescripción por `(semana, día,
    # ejercicio)` porque en la planilla original la semana numeraba el programa
    # entero. Acá reinician en cada bloque, así que la semana 1 día 1 de un
    # bloque choca con la del siguiente: las series de los dos se apilan en la
    # misma prescripción y la segunda muere contra `pset_number_uq`.
    #
    # Se renumeran en vez de cambiar la clave del constructor: la semana relativa
    # que el mesociclo necesita ya la deriva él a partir de las globales, y tocar
    # eso cambiaría también el camino del script.
    globales: dict[tuple[int, int], int] = {}
    for f in filas:
        par = (int(f["Meso #"]), int(f["Semana"]))
        if par not in globales:
            globales[par] = len(globales) + 1
    for f in filas:
        f["Semana"] = globales[(int(f["Meso #"]), int(f["Semana"]))]

    # Y las series se numeran de corrido por ejercicio dentro de la sesión.
    #
    # El mismo ejercicio aparece **dos veces en el mismo día** —una serie pesada
    # arriba y unas de bajada abajo, que es programación normal— y numerando
    # desde 1 en cada aparición las dos chocan contra `pset_number_uq`. En este
    # archivo pasa 79 veces.
    #
    # Las dos apariciones quedan como una sola prescripción con todas sus series.
    # No se pierde nada de lo prescrito: cada serie lleva sus propias
    # repeticiones, su RIR y su carga. Lo que se pierde es que estuvieran
    # escritas en dos renglones, que es una decisión de la planilla y no del plan.
    corridas: dict[tuple[int, int, str], int] = {}
    for f in filas:
        # Nombre propio y no `clave` de nuevo: arriba se usó para (bloque,
        # semana) y son formas distintas.
        cual = (int(f["Semana"]), int(f["Sesión"]), str(f["Ejercicio"]))
        corridas[cual] = corridas.get(cual, 0) + 1
        f["Serie #"] = corridas[cual]

    return filas, "", etiquetas, sorted(set(revisar))


def _numero_de(celdas: list[list[str]], n: int, columna: int, prefijo: str) -> int | None:
    """El número de `Semana 3` o `Sesión 2`, leído en la celda de arriba."""
    if n < 0 or n >= len(celdas):
        return None
    fila = celdas[n]
    if columna >= len(fila):
        return None
    texto = fila[columna]
    if not texto.lower().startswith(prefijo[:5]):
        return None
    nums = _numeros(texto)
    return int(nums[0]) if nums else None


def _series_de(
    cruda: tuple[Any, ...],
    idx: dict[str, int],
    bloque: int,
    semana: int,
    sesion: int,
    hoja: str,
    revisar: list[str],
) -> list[dict[str, Any]]:
    """Un ejercicio como sus series sueltas, que es la forma que espera el constructor."""

    def celda(clave: str) -> Any:
        i = idx.get(clave)
        return cruda[i] if i is not None and i < len(cruda) else None

    nombre = celda("Ejercicio")
    if not nombre or not str(nombre).strip():
        return []

    donde = f"{hoja} · semana {semana} · sesión {sesion} · {str(nombre).strip()}"
    cuantas = _numeros(celda("Series"))
    rmin, rmax = _rango(celda("Reps"), donde, revisar)
    # Medido sobre el archivo: 231 de 231 valores caen entre 0 y 4 y ninguno
    # entre 6 y 10. Es RIR con notación de RPE prestada, no RPE: `@0-1` como RPE
    # querría decir «te sobraban nueve repeticiones», que nadie prescribe.
    rir_min, rir_max = _rango(celda("RIR"), donde, revisar)
    kilos = _numeros(celda("Kg plan"))

    return [
        {
            "Patrón": _patron(list(cruda), idx),
            "Ejercicio": str(nombre).strip(),
            "Básico": "No",
            "Meso #": bloque,
            "Semana": semana,
            "Sesión": sesion,
            "Serie #": numero,
            "Reps plan mín": rmin,
            "Reps plan máx": rmax,
            "RIR plan mín": rir_min,
            "RIR plan máx": rir_max,
            "Kg plan": kilos[0] if kilos else None,
            "Descanso": celda("Descanso"),
            "Observación": celda("Observación"),
            # El histórico del atleta no entra: bajo RLS un entrenador no puede
            # escribir `logged_set`.
            "Reps real": None,
            "Kg real": None,
            "RIR real": None,
            "Comentario": None,
        }
        for numero in range(1, (int(cuantas[0]) if cuantas else 1) + 1)
    ]


def parece_plantilla(xlsx: str) -> bool:
    """Si el libro tiene la forma de esta plantilla.

    Se decide por el encabezado y no por el nombre de las hojas: los entrenadores
    las renombran —`H1`, `BLOQUE 1.3`— y el nombre no dice nada de la estructura.
    """
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    try:
        for hoja in wb.sheetnames:
            for fila in wb[hoja].iter_rows(max_row=30, values_only=True):
                if any(str(c).strip() == ENCABEZADO for c in fila if c):
                    return True
        return False
    finally:
        wb.close()
