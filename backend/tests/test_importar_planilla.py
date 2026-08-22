"""La beta de uno: subir una planilla y que salga una ficha armada.

Corre contra `data/planilla.xlsx`, así que depende de la fixture `seeded` — sin
el archivo estos casos se saltean, igual que el resto de los que miran datos
reales. Es la única planilla con el formato que este importador lee.
"""

from __future__ import annotations

import sqlalchemy as sa


def _habilitar(db) -> None:
    """Prende la bandera del entrenador que hace los pedidos."""
    db.execute(sa.text("RESET ROLE"))
    db.execute(sa.text("UPDATE coach SET puede_importar = true"))
    db.flush()


class TestQuienPuedeImportar:
    def test_sin_la_bandera_no_se_ofrece_ni_se_puede(self, client, seeded, planilla) -> None:
        """Está en prueba: el default es que no, y el 403 lo dice sin prometer nada."""
        r = client.post("/api/athletes/import", files={"archivo": ("p.xlsx", planilla)})
        assert r.status_code == 403, r.text
        assert "prueba" in r.json()["detail"]

    def test_el_perfil_dice_si_ofrecerlo(self, client, seeded, db) -> None:
        """Lo que hace que la interfaz no dibuje un botón que contesta 403."""
        assert client.get("/api/coach").json()["puede_importar"] is False
        _habilitar(db)
        assert client.get("/api/coach").json()["puede_importar"] is True


class TestLoQueDeja:
    def test_crea_una_ficha_nueva_con_su_programa(self, client, seeded, db, planilla) -> None:
        antes = len(client.get("/api/athletes").json())
        _habilitar(db)

        r = client.post("/api/athletes/import", files={"archivo": ("p.xlsx", planilla)})
        assert r.status_code == 201, r.text
        cuerpo = r.json()

        assert cuerpo["athlete_name"], "el nombre sale de la celda B5 de la planilla"
        assert cuerpo["creados"]["mesocycles"] >= 1
        assert cuerpo["creados"]["prescribed_sets"] >= 1
        assert len(client.get("/api/athletes").json()) == antes + 1

    def test_no_trae_el_historial_del_atleta(self, client, seeded, db, planilla) -> None:
        """El caso que justifica que esto sea un endpoint y no el script.

        Bajo RLS un entrenador no puede escribir `logged_set`: registrar es el
        acto del atleta. Si esto importara el historial, la base rechazaría la
        transacción entera — y si algún día dejara de rechazarla, un entrenador
        estaría cargando lo que otro «hizo», que es por donde entran los datos de
        entrenamiento inventados.
        """
        _habilitar(db)
        r = client.post("/api/athletes/import", files={"archivo": ("p.xlsx", planilla)})
        assert r.status_code == 201, r.text

        # Por id y no por nombre: el atleta sembrado se llama igual —los dos
        # salen de la misma celda B5— así que contar por nombre suma las 1.199
        # series del otro y el test falla diciendo lo contrario de lo que pasa.
        db.execute(sa.text("RESET ROLE"))
        registros = db.execute(
            sa.text("SELECT count(*) FROM logged_set WHERE athlete_id = :id"),
            {"id": r.json()["athlete_id"]},
        ).scalar()
        assert registros == 0, "la importación no debe traer series registradas"
        assert "logged_sets" not in r.json()["creados"]

    def test_lo_que_no_se_pudo_desambiguar_vuelve_para_revisar(
        self, client, seeded, db, planilla
    ) -> None:
        """No se inventa: se deja nulo y se avisa. Es la regla del parser original."""
        _habilitar(db)
        cuerpo = client.post("/api/athletes/import", files={"archivo": ("p.xlsx", planilla)}).json()
        assert isinstance(cuerpo["revisar"], list)


class TestLoQueRechaza:
    def test_un_archivo_que_no_es_planilla(self, client, seeded, db) -> None:
        _habilitar(db)
        r = client.post("/api/athletes/import", files={"archivo": ("notas.txt", b"hola")})
        assert r.status_code == 415, r.text

    def test_un_xlsx_sin_las_hojas_que_hacen_falta(self, client, seeded, db) -> None:
        """Un Excel cualquiera no es esta planilla, y decirlo es mejor que un 500."""
        import io

        import openpyxl

        _habilitar(db)
        wb = openpyxl.Workbook()
        wb.active.title = "Hoja1"
        buf = io.BytesIO()
        wb.save(buf)

        r = client.post("/api/athletes/import", files={"archivo": ("otra.xlsx", buf.getvalue())})
        assert r.status_code == 422, r.text
        assert "DATOS" in r.json()["detail"]


class TestLaPlantillaDeLosEntrenadores:
    """El otro formato: la plantilla que los entrenadores usan de verdad.

    No tiene hoja `DATOS` ni `ATLETA`. Es una **grilla**: las columnas son
    semanas —el encabezado se repite hacia la derecha— y las filas son sesiones.
    Leerla como una lista trae un quinto del archivo y hace creer que cada bloque
    tiene una sola semana.
    """

    def test_lee_la_grilla_entera_y_no_la_primera_columna(self) -> None:
        """El error que se cometió escribiendo esto, hecho test.

        Se construye una hoja con dos semanas a lo ancho. Leyendo sólo la primera
        columna salen la mitad de las series y una sola semana, y todo lo demás
        parece andar — que es por qué hace falta este caso.
        """
        import io

        import openpyxl

        from importer.plantilla import leer

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BLOQUE 1"
        cabecera = ["Patron mov.", "Ejercicio", "Series", "Repeticiones @RPE", "RIR", "Kilos"]
        # Semana 1 desde la columna B, semana 2 desde la I.
        ws.cell(row=1, column=2, value="Semana 1")
        ws.cell(row=1, column=9, value="Semana 2")
        ws.cell(row=2, column=2, value="Sesión 1")
        ws.cell(row=2, column=9, value="Sesión 1")
        for d, texto in enumerate(cabecera):
            ws.cell(row=3, column=2 + d, value=texto)
            ws.cell(row=3, column=9 + d, value=texto)
        for d, valor in enumerate(["SQUAT", "HIGH BAR SQUAT", 3, "6 a 8", "@3", "80k"]):
            ws.cell(row=4, column=2 + d, value=valor)
        for d, valor in enumerate(["SQUAT", "HIGH BAR SQUAT", 3, "6 a 8", "@2", "90k"]):
            ws.cell(row=4, column=9 + d, value=valor)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            tmp.write(buf.getvalue())
            tmp.flush()
            filas, _, etiquetas, _avisos = leer(tmp.name)

        assert etiquetas == {1: "BLOQUE 1"}
        assert sorted({f["Semana"] for f in filas}) == [1, 2], "las columnas son semanas"
        # Tres series por semana, dos semanas.
        assert len(filas) == 6

        semana2 = [f for f in filas if f["Semana"] == 2]
        assert semana2[0]["RIR plan mín"] == 2, "la semana 2 aprieta el RIR"
        assert semana2[0]["Kg plan"] == 90

    def test_el_rir_con_arroba_es_rir_y_no_rpe(self) -> None:
        """`@2-3` es RIR, medido: 231 de 231 valores del archivo real caen entre 0
        y 4, y ninguno entre 6 y 10.

        Leerlo como RPE invertiría la señal que es el corazón del producto: RPE 8
        equivale a RIR 2, así que la aplicación diría que el atleta entrena cerca
        del fallo cuando entrena lejos.
        """
        from importer.plantilla import _rango

        revisar: list[str] = []
        assert _rango("@2-3", "x", revisar) == (2.0, 3.0)
        assert _rango("@0-1", "x", revisar) == (0.0, 1.0)
        assert revisar == []

    def test_una_fecha_no_se_reconstruye_en_repeticiones(self) -> None:
        """Excel convierte `4-5` en una fecha al tipearlo. La intención original no
        se puede recuperar, así que queda en nulo y se marca — nunca se deduce del
        mes y el día."""
        from datetime import datetime

        from importer.plantilla import _rango

        revisar: list[str] = []
        assert _rango(datetime(2026, 5, 4), "sesión 1 · Sentadilla", revisar) == (None, None)
        assert revisar == ["sesión 1 · Sentadilla: Excel convirtió el valor en una fecha"]
